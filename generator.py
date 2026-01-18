#!/usr/bin/env python3

"""
Label generator:
- Reads spreadsheets (CSV, Excel ".xlsx", Apple Numbers ".numbers", OpenDocument ".ods") with columns:
  - name
  - description
  - top_symbol
  - side_symbol
  - reorder_url
- Generates an SVG label per row by filling a template SVG (template.svg)
- Creates QR code SVG using segno and embeds it into the template
- Converts final SVG into output format (one file per label) using cairosvg
"""

import argparse
import io
from pathlib import Path
import re
import sys
from urllib.parse import quote
import segno
from jinja2 import Template
import cairosvg
import colorama
import requests

import shapes

# Configuration
LABEL_WIDTH_MM = 36
LABEL_HEIGHT_MM = 7.7  # Nominally 9mm, but we can't print all the way to the edges
LABEL_PRINTER_DPI = 150  # for Brother P710BT label printer


# simple logging functions
# enum representing log levels
class LogLevel:
    """Enum representing log levels for the label generator."""

    DEBUG = -2
    INFO = -1
    NORMAL = 0
    WARNING = 1
    ERROR = 2


LOG_LEVEL = LogLevel.NORMAL
xprint = print  # backup the normal print function


def debug(msg: str, *args, **kwargs) -> None:
    """Log a debug message if the current log level is DEBUG or lower."""
    if LOG_LEVEL <= LogLevel.DEBUG:
        xprint(
            f"{colorama.Fore.BLUE}[DEBUG] {msg}{colorama.Style.RESET_ALL}",
            *args,
            **kwargs,
        )


def info(msg: str, *args, **kwargs) -> None:
    """Log an info message if the current log level is INFO or lower."""
    if LOG_LEVEL <= LogLevel.INFO:
        xprint(
            f"{colorama.Fore.GREEN}[INFO ] {msg}{colorama.Style.RESET_ALL}",
            *args,
            **kwargs,
        )


def print(msg: str, *args, **kwargs) -> None:
    """Log a normal message if the current log level is NORMAL or lower."""
    if LOG_LEVEL <= LogLevel.NORMAL:
        xprint(msg, *args, **kwargs)


def warn(msg: str, *args, **kwargs) -> None:
    """Log a warning message if the current log level is WARNING or lower."""
    if LOG_LEVEL <= LogLevel.WARNING:
        xprint(
            f"{colorama.Fore.YELLOW}[WARN ] {msg}{colorama.Style.RESET_ALL}",
            *args,
            **kwargs,
        )


def error(msg: str, *args, **kwargs) -> None:
    """Log an error message if the current log level is ERROR or lower."""
    if LOG_LEVEL <= LogLevel.ERROR:
        if "file" not in kwargs:
            kwargs["file"] = sys.stderr
        xprint(
            f"{colorama.Fore.RED}[ERROR] {msg}{colorama.Style.RESET_ALL}",
            *args,
            **kwargs,
        )


#######################
# Program code
#######################


def read_template(template_file: Path) -> Template:
    """Read an SVG template file and return a Template object."""
    with template_file.open("r", encoding="utf-8") as f:
        return Template(f.read())


def shorten_url(url: str) -> str:
    """Shorten a URL using the v.gd API."""
    if not url.startswith("http://") and not url.startswith("https://"):
        return url
    api_base = "https://v.gd/create.php?format=simple&url="
    try:
        response = requests.get(api_base + quote(url, safe=""), timeout=5)
        response.raise_for_status()
        payload = response.text.strip()

        # strip scheme for more compact encoding
        if payload.startswith("http://"):
            payload = payload[len("http://") :]
        elif payload.startswith("https://"):
            payload = payload[len("https://") :]

        return payload
    except requests.RequestException as e:
        warn(f"URL shortening failed for {url}: {e}")
        return url


def make_qr_svg(content: str, scale_mm: float, qr_type: str = "micro") -> tuple[str, float]:
    """Generate a QR code as an SVG string and return it along with its size in modules."""
    # Create QR as SVG string. segno uses px units; we'll embed SVG and scale via width/height attrs
    # Try running the URL through a shortener, then remove the leading 'http(s)://', so that micro QR can fit more data

    if not content:
        return "", 0

    if qr_type not in ("micro", "standard"):
        raise ValueError(f"Invalid qr_type: {qr_type}")

    payload = content
    debug(f"Generating {qr_type} QR for content: {content}")
    qr = None
    if qr_type == "standard":
        qr = segno.make(content)
    elif qr_type == "micro":
        payload = shorten_url(content)
        try:
            # Try to make a micro QR code
            qr = segno.make(payload, micro=(qr_type == "micro"))
        except ValueError:
            # Fallback to standard QR if micro QR is not possible
            info(f"URL too long for micro QR, using standard QR: {content}")
            qr = segno.make(payload, micro=False)
    if qr is None:
        raise RuntimeError("Failed to generate QR code")

    buf = io.BytesIO()
    qr.save(buf, scale=scale_mm, kind="svg", border=0)  # raw svg bytes
    svg_bytes = buf.getvalue().decode("utf-8")
    # strip xml declaration to avoid duplicates when embedding
    svg_body = "\n".join(line for line in svg_bytes.splitlines() if not line.startswith("<?xml"))
    return (svg_body, qr.symbol_size()[0])


def mm_to_px(mm: float) -> int:
    """Convert millimeters to pixels based on the label printer DPI."""
    inches = mm / 25.4
    return int(inches * LABEL_PRINTER_DPI)


######################
# Spreadsheet Handling
######################

# Required columns for the parts CSV/Excel/Numbers files
REQUIRED_COLUMNS = ["name", "description"]
OPTIONAL_COLUMNS = ["top_symbol", "side_symbol", "reorder_url", "top_icon", "side_icon", "qr_svg", "label"]
IMAGE_COLUMNS = ["top_icon", "side_icon", "qr_svg", "label"]


def parse_csv(csv_file: Path, delimiter: str = ",") -> list[dict[str, str]]:
    """Parse a CSV file and return a list of dictionaries representing each row."""
    from csv import DictReader

    rows = []
    with open(csv_file, newline="", encoding="utf-8") as f:
        reader = DictReader(f, delimiter=delimiter)
        headers = reader.fieldnames
        if headers is None:
            raise ValueError(f"No headers found in CSV file: {csv_file}")
        missing_columns = [col for col in REQUIRED_COLUMNS if col not in headers]
        if missing_columns:
            raise ValueError(f"Missing required columns in CSV file {csv_file}: {missing_columns}")
        for row in reader:
            rows.append(
                {
                    "name": row.get("name", "").strip(),
                    "description": row.get("description", "").strip(),
                    "top_symbol": row.get("top_symbol", "").strip(),
                    "side_symbol": row.get("side_symbol", "").strip(),
                    "reorder_url": row.get("reorder_url", "").strip(),
                }
            )
    return rows


def write_csv(rows: list[dict[str, str]], output_file: Path, delimiter: str = ",") -> None:
    """Write a list of dictionaries to a CSV file."""
    from csv import DictWriter

    if not rows:
        raise ValueError("No data to write")

    headers = list(rows[0].keys())
    with output_file.open("w", newline="", encoding="utf-8") as f:
        writer = DictWriter(f, fieldnames=headers, delimiter=delimiter)
        writer.writeheader()
        for row in rows:
            writer.writerow(row)


def parse_excel(excel_file: Path) -> list[dict[str, str]]:
    """Parse an Excel file and return a list of dictionaries representing each row."""
    from openpyxl import load_workbook

    rows = []
    wb = load_workbook(excel_file, data_only=True)
    ws = wb.active
    headers = [cell.value for cell in ws[1]]
    if headers is None:
        raise ValueError(f"No headers found in Excel file: {excel_file}")
    missing_columns = [col for col in REQUIRED_COLUMNS if col not in headers]
    if missing_columns:
        raise ValueError(f"Missing required columns in Excel file {excel_file}: {missing_columns}")

    for row in ws.iter_rows(min_row=2, values_only=True):
        row_dict = {header: (value if value is not None else "") for header, value in zip(headers, row)}
        rows.append(
            {
                "name": row_dict.get("name", "").strip(),
                "description": row_dict.get("description", "").strip(),
                "top_symbol": row_dict.get("top_symbol", "").strip(),
                "side_symbol": row_dict.get("side_symbol", "").strip(),
                "reorder_url": row_dict.get("reorder_url", "").strip(),
            }
        )
    return rows


def write_excel(rows: list[dict[str, str]], output_file: Path) -> None:
    """Write a list of dictionaries to an Excel file."""
    from openpyxl import Workbook, drawing

    if not rows:
        raise ValueError("No data to write")
    wb = Workbook()
    ws = wb.active
    ws.append(list(rows[0].keys()))
    for row in rows:
        ws.append(list(row.values()))
    wb.save(output_file)


def parse_numbers(numbers_file: Path) -> list[dict[str, str]]:
    """Parse a Numbers file and return a list of dictionaries representing each row."""
    from numbers_parser import Document

    doc = Document(numbers_file)
    sheet = doc.sheets[0]
    table = sheet.tables[0]
    headers = [cell.value for cell in table.rows()[0]]
    if headers is None:
        raise ValueError(f"No headers found in Numbers file: {numbers_file}")
    missing_columns = [col for col in REQUIRED_COLUMNS if col not in headers]
    if missing_columns:
        raise ValueError(f"Missing required columns in Numbers file {numbers_file}: {missing_columns}")

    rows = []
    for row in table.rows()[1:]:
        row_dict = {
            header: (cell.value if cell is not None and cell.value is not None else "")
            for header, cell in zip(headers, row)
        }
        rows.append(
            {
                "name": row_dict.get("name", "").strip(),
                "description": row_dict.get("description", "").strip(),
                "top_symbol": row_dict.get("top_symbol", "").strip(),
                "side_symbol": row_dict.get("side_symbol", "").strip(),
                "reorder_url": row_dict.get("reorder_url", "").strip(),
            }
        )
    return rows


def write_numbers(rows: list[dict[str, str]], output_file: Path) -> None:
    """Write a list of dictionaries to a Numbers file."""
    from numbers_parser import Document

    doc = Document()
    sheet = doc.sheets[0]
    table = sheet.tables[0]
    headers = [cell.value for cell in table.rows()[0]]
    for row in rows:
        table.append_row([row.get(header, "") for header in headers])
    doc.save(output_file)


def parse_ods(ods_file: Path) -> list[dict[str, str]]:
    """Parse an ODS file and return a list of dictionaries representing each row."""
    from ezodf import opendoc

    doc = opendoc(ods_file)
    sheet = doc.sheets[0]
    headers = [cell.value for cell in sheet.rows()[0]]
    if headers is None:
        raise ValueError(f"No headers found in ODS file: {ods_file}")
    missing_columns = [col for col in REQUIRED_COLUMNS if col not in headers]
    if missing_columns:
        raise ValueError(f"Missing required columns in ODS file {ods_file}: {missing_columns}")

    rows = []
    for row in sheet.rows()[1:]:
        row_dict = {header: (cell.value if cell.value is not None else "") for header, cell in zip(headers, row)}
        rows.append(
            {
                "name": row_dict.get("name", "").strip(),
                "description": row_dict.get("description", "").strip(),
                "top_symbol": row_dict.get("top_symbol", "").strip(),
                "side_symbol": row_dict.get("side_symbol", "").strip(),
                "reorder_url": row_dict.get("reorder_url", "").strip(),
            }
        )
    return rows


def write_ods(rows: list[dict[str, str]], output_file: Path) -> None:
    """Write a list of dictionaries to an ODS file."""
    from ezodf import newdoc, Sheet

    if not rows:
        raise ValueError("No data to write")
    doc = newdoc(doctype="ods", filename=output_file)
    sheet = Sheet("Sheet1", size=(len(rows) + 1, len(rows[0])))
    doc.sheets += sheet
    headers = list(rows[0].keys())
    for col, header in enumerate(headers):
        sheet[0, col].set_value(header)
    for row_idx, row in enumerate(rows, start=1):
        for col_idx, header in enumerate(headers):
            sheet[row_idx, col_idx].set_value(row.get(header, ""))
    doc.save()


def parse_spreadsheet(spreadsheet_file: Path) -> list[dict[str, str]]:
    """Parse a spreadsheet file (CSV, TSV, Excel, ODS, Numbers) and return a list of dictionaries representing each row."""
    extension = spreadsheet_file.suffix.lower()
    rows = []
    match extension:
        case ".csv":
            rows = parse_csv(spreadsheet_file, delimiter=",")
        case ".tsv":
            rows = parse_csv(spreadsheet_file, delimiter="\t")
        case ".xls" | ".xlsx":
            rows = parse_excel(spreadsheet_file)
        case ".ods":
            rows = parse_ods(spreadsheet_file)
        case ".numbers":
            rows = parse_numbers(spreadsheet_file)
        case _:
            raise ValueError(f"Unsupported spreadsheet format: {spreadsheet_file.suffix}")
    debug(f"Parsed {len(rows)} rows from {spreadsheet_file}")
    return rows


def write_spreadsheet(rows: list[dict[str, str]], output_file: Path, delimiter: str = ",") -> None:
    """Write a list of dictionaries to a spreadsheet file."""
    if not rows:
        raise ValueError("No data to write")
    debug(f"Writing {len(rows)} rows to {output_file}")

    extension = output_file.suffix.lower()
    match extension:
        case ".csv":
            write_csv(rows, output_file, delimiter=",")
        case ".tsv":
            write_csv(rows, output_file, delimiter="\t")
        case ".xlsx":
            write_excel(rows, output_file)
        case ".ods":
            write_ods(rows, output_file)
        case ".numbers":
            write_numbers(rows, output_file)
        case _:
            raise ValueError(f"Unsupported spreadsheet format: {output_file.suffix}")


def sanitize_svg(svg: str) -> str:
    """Validate and clean up an SVG string."""
    # Allow empty strings (for icons with no generator)
    if not svg:
        return ""

    # Remove XML declaration if present
    svg = re.sub(r"<\?xml.*?\?>", "", svg).strip()
    # Remove DOCTYPE declaration if present
    svg = re.sub(r"<!DOCTYPE.*?>", "", svg).strip()
    # Remove comments
    svg = re.sub(r"<!--.*?-->", "", svg, flags=re.DOTALL).strip()
    # Remove unnecessary whitespace between tags
    svg = re.sub(r">\s+<", "><", svg).strip()
    # Remove leading and trailing whitespace
    svg = svg.strip()

    # If content remains, ensure the SVG is valid (must start and end with angle brackets)
    if svg and not (svg.startswith("<") and svg.endswith(">")):
        raise ValueError("Invalid SVG content")

    return svg


def generate_labels(
    parts_file_path: Path,
    template_file: Path,
    output_dir: Path,
    qr_type: str = "micro",
    output_format: str = "png",
) -> list[dict[str, str]]:
    """Generate labels for the parts listed in the spreadsheet using the provided template."""

    template = read_template(template_file)
    rows = parse_spreadsheet(parts_file_path)
    output_dir.mkdir(parents=True, exist_ok=True)

    for row in rows:
        name = row["name"]
        description = row["description"]
        top_symbol = row["top_symbol"].lower()
        side_symbol = row["side_symbol"].lower()
        reorder_url = row["reorder_url"]

        # Generate icons, if not provided
        if not row.get("top_icon"):
            top_icon = ""
            if top_symbol:
                top_gen = shapes.IconRegistry.top_generators.get(top_symbol)
                if top_gen is None:
                    error(f"top icon generator not found for '{top_symbol}'")
                else:
                    top_icon = top_gen()
            if top_icon:
                row["top_icon"] = sanitize_svg(top_icon)
                debug(f"Generated top icon for {name}+{description}")
            else:
                row["top_icon"] = ""
        else:
            top_icon = sanitize_svg(row["top_icon"])
            # strip xml declaration
            top_icon = re.sub(r"<\?xml.*?\?>", "", top_icon).strip()
            debug(f"Using existing top icon for {name}+{description}")

        if not row.get("side_icon"):
            side_icon = ""
            if side_symbol:
                side_gen = shapes.IconRegistry.side_generators.get(side_symbol)
                if side_gen is None:
                    error(f"side icon generator not found for '{side_symbol}'")
                else:
                    side_icon = side_gen()
            if side_icon:
                row["side_icon"] = sanitize_svg(side_icon)
                debug(f"Generated side icon for {name}+{description}")
            else:
                row["side_icon"] = ""
        else:
            side_icon = sanitize_svg(row["side_icon"])
            debug(f"Using existing side icon for {name}+{description}")

        if not row.get("qr_svg"):
            qr_svg, qr_width = make_qr_svg(reorder_url, LABEL_HEIGHT_MM, qr_type=qr_type)
            row["qr_svg"] = qr_svg
            debug(f"Generated QR code for {name}+{description}")
        else:
            qr_svg = sanitize_svg(row["qr_svg"])
            match = re.search(r'width="([\d.]+)mm"', qr_svg)
            qr_width = float(match.group(1)) if match else LABEL_HEIGHT_MM
            debug(f"Using existing QR code for {name}+{description}")

        svg_filled = sanitize_svg(
            template.render(
                {
                    "LABEL_WIDTH_MM": LABEL_WIDTH_MM,
                    "LABEL_HEIGHT_MM": LABEL_HEIGHT_MM,
                    "name": name,
                    "description": description,
                    "top_icon_svg": top_icon,
                    "side_icon_svg": side_icon,
                    "qr_svg": qr_svg,
                    "qr_size": qr_width,  # size of the generated QR code in mm
                }
            )
        )
        row["label"] = svg_filled

        # Use name for filename (sanitize for filesystem)
        filename = name.replace("/", "-").replace(" ", "_") + "+" + description.replace(" ", "_").replace("/", "-")
        out_file_path = output_dir / f"{filename}.{output_format}"

        match output_format:
            case "svg":
                # write SVG output directly (no conversion needed)
                with out_file_path.open("w", encoding="utf-8") as f:
                    f.write(svg_filled)
            case "pdf":
                # Write PDF output
                cairosvg.svg2pdf(
                    bytestring=svg_filled.encode("utf-8"),
                    write_to=str(out_file_path),
                    output_width=mm_to_px(LABEL_WIDTH_MM),
                    output_height=mm_to_px(LABEL_HEIGHT_MM),
                )
            case "png":
                # Write PNG output
                cairosvg.svg2png(
                    bytestring=svg_filled.encode("utf-8"),
                    write_to=str(out_file_path),
                    output_width=mm_to_px(LABEL_WIDTH_MM),
                    output_height=mm_to_px(LABEL_HEIGHT_MM),
                )
        print(f"Generated: {out_file_path}")
    return rows


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Generate labels for Gridfinity bins from a spreadsheet")
    parser.add_argument(
        "parts_file",
        type=Path,
        nargs="?",
        default=Path("parts.csv"),
        help="Path to the parts CSV file (default: parts.csv)",
    )
    parser.add_argument(
        "template_file",
        type=Path,
        nargs="?",
        default=Path("template.svg"),
        help="Path to the template SVG file (default: template.svg)",
    )
    parser.add_argument(
        "output_dir",
        type=Path,
        nargs="?",
        default=Path("output"),
        help="Path to the output directory (default: output)",
    )
    parser.add_argument(
        "--format",
        choices=["pdf", "png", "svg"],
        default="png",
        help="Output file format (default: png)",
    )
    parser.add_argument(
        "--qr-type",
        choices=["micro", "standard"],
        default="micro",
        help="Type of QR code to generate (default: micro)",
    )
    parser.add_argument(
        "-q",
        "--quiet",
        action="store_true",
        help="Suppress non-error log messages",
    )
    parser.add_argument(
        "-v",
        "--verbose",
        action="count",
        default=0,
        help="Increase logging verbosity (use multiple times for more verbose: -v, -vv, -vvv)",
    )
    parser.add_argument(
        "--export",
        default=False,
        action="store_true",
        help="Export (update) the parts list back to the spreadsheet",
    )

    args = parser.parse_args()

    # Set log level based on verbosity/quiet flags
    if args.quiet:
        LOG_LEVEL = LogLevel.ERROR  # only errors
    else:
        LOG_LEVEL = max(LogLevel.DEBUG, LogLevel.NORMAL - args.verbose)  # decrease log level with more -v

    # Check input files
    if not args.parts_file.exists():
        error(f"Parts file '{args.parts_file}' does not exist.")
        exit(1)
    if not args.template_file.exists():
        error(f"Template file '{args.template_file}' does not exist.")
        exit(1)

    rows = generate_labels(
        args.parts_file,
        args.template_file,
        args.output_dir,
        qr_type=args.qr_type,
        output_format=args.format,
    )

    if args.export:
        write_spreadsheet(rows, args.parts_file)
